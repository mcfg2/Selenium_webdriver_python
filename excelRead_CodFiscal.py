from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as E
from openpyxl.styles import PatternFill
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import  Select
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as O

service_obj = Service("PATH")  #path chromedriver
driver = webdriver.Chrome(service=service_obj)
driver.maximize_window()
URL = "URL"  #URL webpage
Excel_file = "PATH"  #path excel file
Excel_worksheet = "Raport"  #excel worksheet name


driver.get(URL)
driver.find_element(By.XPATH, "//input[@placeholder='Username']").send_keys("****")  #login
driver.find_element(By.XPATH, "//input[@placeholder='Parola']").send_keys("****")  #password
driver.find_element(By.XPATH, "//button[normalize-space()='LOG IN']").click()
time.sleep(2)
driver.find_element(By.XPATH, "(//div[@class='modules-item__icon-block'])[1]").click()

wb = O.load_workbook(Excel_file)
ws = wb[Excel_worksheet]

for r in range(2,ws.max_row +1):
    d = str(ws.cell(r, 1).value)
    driver.find_element(By.XPATH, "//input[@id='search']").send_keys(d)
    driver.find_element(By.XPATH, "//input[@id='search']").send_keys(Keys.ENTER)
    time.sleep(25)
    driver.find_element(By.XPATH, "(//i[@class='material-icons more'][normalize-space()='expand_more'])[1]").click()
    time.sleep(1)
    e = str(ws.cell(r,2).value)
    if str(e) in driver.find_element(By.XPATH, "//div[5]//div[1]//div[1]//div[1]").text:
        ws.cell(r,3).value = "Pass"
        ws.cell(r,3).fill = PatternFill("solid", start_color="009900")
    else:
        ws.cell(r,3).value = "Fail"
        ws.cell(r, 3).fill = PatternFill("solid", start_color="ff0000")
    wb.save(Excel_file)
    wb.close()
